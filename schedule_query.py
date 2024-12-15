from schedule_helper import *
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.styles import Alignment

import sys
from subprocess import Popen
import signal

def exit_handler(signum,frame):
    print("\nExiting...")
    print("Process exit with code " + str(signum))
    sys.exit(0)

def index_to_time(N):
    suffix = 'AM'
    if N >= 10:
        time = str((N-8)//2)
        suffix = 'PM'
    else:
        time = str((N + 16)//2)
    if N%2:
        time += f":30{suffix}"
    else:
        time += f":00{suffix}"
    return time

def tuple_to_time(start, end):
    if start > 13: start -= 12
    if end > 13: end -= 12
    start_time = str(int(start//1))
    end_time = str(int(end//1))
    if 10*start%2:
        start_time += ':30'
    else:
        start_time += ':00'
    if 10%end%2:
        end_time += ':30'
    else:
        end_time += ':00'
    return f"{start_time}-{end_time}"
    

def write_sch_to_excel(wb:Workbook, ws, schedule, sch_matrix: list):
    legend = list(schedule)
    thin_border = Border(
        left=Side(border_style=BORDER_THIN, color='D8D8D8'),
        right=Side(border_style=BORDER_THIN, color='D8D8D8'),
        top=Side(border_style=BORDER_THIN, color='D8D8D8'),
        bottom=Side(border_style=BORDER_THIN, color='D8D8D8')
    )
    colors = {
        0: "FFFFFF",  # White
        1: "6EFFA8",  # Green
        2: "70C3FF",  # Blue
        3: "FF7070",  # Red
        4: "FFC670",  # Yellow
        5: "707EFF",  # Dark Blue
        6: "FF70BF"   # Pink
    }
    for i in range(len(DAYS)):
        cell = ws.cell(row=1, column=i+2)
        cell.value = DAYS[i][:3]
        cell.alignment = Alignment(horizontal="center")
    
    for i in range(len(sch_matrix)):
        cell = ws.cell(row=i+2, column=1)
        time = index_to_time(i)
        cell.value = time
        for j in range(len(sch_matrix[i])):
            cell = ws.cell(row=i+2, column=j+2)
            value = sch_matrix[i][j]
            if value != 0:
                cell.fill = PatternFill(start_color=colors[value], end_color=colors[value], fill_type="solid")
            else: # Border for empty cells only
                cell.border = thin_border
            cell.value = ""
    # sch_matrix_T = matrix_transpose(sch_matrix)

    # # TODO: Text on schedule
    # for i in range(len(sch_matrix_T)):
    #     day = DAYS[i]
    #     start = 0
    #     end = 0
    #     val = sch_matrix_T[0]
    #     for j in range(len(sch_matrix_T)):
    #         if sch_matrix_T[j] != 0:
    #             start = j
    #             val = sch_matrix_T[j]
    #         if start != 0 and sch_matrix_T[j] == 0:
    #             end = j-1
    #             mid = (start+end//2)+(1-(end-start)%2)
    #             course = list(schedule)[val]
    #             section = schedule[course]
    #             for days in section:
    #                 if days == day:
    #                     cell = ws.cell(column=mid+2, row=j+2)
    #                     cell.value = course[:4]
    #                     cell = ws.cell(column=mid+3, row=j+2)
    #                     time = tuple_to_time(day[0], day[1])
    #                     cell.value = time
    #                     print(time)
    #             start, end = 0, 0
                        

            
    for i in range(len(legend)):
        cell = ws.cell(row=i+4, column=9)
        cell.fill = PatternFill(start_color=colors[i+1], end_color=colors[i+1], fill_type="solid")
        cell = ws.cell(row=i+4, column=10)
        section_no = list(schedule[legend[i]])[0]
        cell.value = f"{legend[i]} - {section_no}"
    wb.save("schedule_timings.xlsx")


def wildcard_query_multiple(comb_list,comb_tuple):
    print("\nYOU ARE NOW IN QUERY MODE. Try querying possible combinations (like '1 1 1 *')")
    repeat = True
    while repeat:
        wildcard_str = input("Enter course options: ").split(' ')
        if len(wildcard_str) == 1 and wildcard_str[0] == '':
            sys.exit()
        sch_tuple = []
        # if len(sch_tuple) != len(comb_tuple[0]):
        #     print("You have not entered your combination acc. to the correct amount of classes")
        #     continue
        num_count = 0
        for i in range(len(wildcard_str)):
            if wildcard_str[i] == '*':
                sch_tuple.append(wildcard_str[i])
            else:
                num_count += 1
                sch_tuple.append(int(wildcard_str[i]))
        # If compatible schedules are found
        if len(get_schedules_for_query(sch_tuple,comb_list,comb_tuple)) > 0:
            repeat = False
            break
        else:
            print("The combination you entered above is not a valid combination. Please try again in query mode.")


def get_schedules_for_query(sch_query_tuple,comb_list,comb_tuple):
    queried_schedules = []
    num_count = len(sch_query_tuple) - sch_query_tuple.count('*')
    for combination in comb_tuple:
        count = 0
        for i in range(len(combination)):
            if sch_query_tuple[i] != '*' and combination[i] == sch_query_tuple[i]:
                count += 1
            if count == num_count:
                if len(queried_schedules) == 0:
                    print("\nHere are the combinations that work: ")
                print(combination)
                queried_schedules.append((comb_list[comb_tuple.index(combination)],combination))
                break   # New addition, test
    return queried_schedules


def depreciated_wildcard_query(comb_tuple):
    repeat = True
    while repeat:
        print("This combination is not compatible.")
        print("Enter the wildcard * in one of the digits to get possible combinations like \"3 6 * 9\"")
        print("Leave input blank to exit")
        wildcard_str = input("Enter course options: ").split(' ')
        if len(wildcard_str) == 1 and wildcard_str[0] == '':
            sys.exit()
        sch_tuple = []
        # wildcard_start, wildcard_end = -1, -1
        wildcard_pos = -1
        for i in range(len(wildcard_str)):
            if wildcard_str[i] == "*":
                wildcard_pos = i
                sch_tuple.append(wildcard_str[i])
            # elif wildcard_start > 0 and wildcard_end < 0:
            #     wildcard_end = i
            #     sch_tuple.append(int(wildcard_str[i]))
            else:
                sch_tuple.append(int(wildcard_str[i]))
        sch_tuple = tuple(sch_tuple)
        if wildcard_pos > 0:
            print("Here are the following combinations that work: ")
            # middle_tuple = []
            # for i in range(wildcard_end-wildcard_start):
            #     middle_tuple.append(list(range(1,20)))
            # cartesian = itertools.product(*middle_tuple)
            # for i in cartesian:
            #     tuple_i = []
            #     tuple_i.extend(sch_tuple[:wildcard_start])
            #     tuple_i.extend(list(i))
            #     tuple_i.extend(sch_tuple[wildcard_end:])
            #     tuple_i = tuple(tuple_i)
            #     if tuple_i in comb_tuple: 
            #         repeat = False
            #         print(tuple_i)
            for i in range(1,20):
                tuple_i = []
                tuple_i.extend(sch_tuple[:wildcard_pos])
                tuple_i.append(i)
                tuple_i.extend(sch_tuple[wildcard_pos+1:])
                tuple_i = tuple(tuple_i)
                if tuple_i in comb_tuple: 
                    #######repeat = False
                    print(tuple_i)


def open_excel_sheet(program: str) -> None:
    try:
        # TODO Possibly include stdout=subprocess.PIPE, stderr=subprocess.STDOUT to suppress errors?
        process = Popen([program, "schedule_timings.xlsx"])
    except FileNotFoundError as fe:
        print("\nProgram not found. Please provide proper path")
        print(fe)
    except Exception as e:
        print("Something went wrong")
        print(e)


def main_schedule_query(schedules, comb_list: list, comb_tuple: list[tuple]) -> bool:
    print("Query particular schedules from txt file")
    print("Separate input of course options by spaces (like \"1 1 1 1\")")
    in_str = input("Enter course options: ")
    if in_str == '':
        wildcard_query_multiple(comb_list,comb_tuple)
        sys.exit()
    else:
        in_str = in_str.split(' ')
    sch_tuple = []
    for i in range(len(in_str)):
        sch_tuple.append(int(in_str[i]))
    sch_tuple = tuple(sch_tuple)

    if sch_tuple in comb_tuple:
        schedule = comb_list[comb_tuple.index(sch_tuple)]
        sch_matrix = get_schedule_matrix(schedule)
        print("Selected schedule: ")
        print(schedule)
        # print(sch_matrix)
        wb = Workbook()
        ws = wb.active
        write_sch_to_excel(wb,ws,schedule,sch_matrix)
        # if openWith != "":
        #     open_excel_sheet(program=openWith)       
        return True
    else:
        print("\nThat schedule is not compatible. Entering query mode...")
        wildcard_query_multiple(comb_list,comb_tuple)
        return False

signal.signal(signal.SIGINT, exit_handler)

if __name__ == "__main__":

    schedules = get_schedule_as_dict()
    comb_list, comb_tuple = get_compatible_combinations(schedules)

    main_schedule_query(schedules, comb_list, comb_tuple)
